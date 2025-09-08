import openpyxl
import urllib
import uuid
import os
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.core.paginator import Paginator, EmptyPage
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.urls import reverse
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.files.storage import default_storage

from parametre.models import TypeCours, Quartier, Tribu, Departement
from session.forms import SessionForm, CertificatForm, CoursForm, CheminantForm, InscriptionForm
from session.models import Session, Certificat, Cours, Inscription, Question, Reponse
from shared.enum import SessionStatut, SituationMatrimoniale, Genre, StatutCertificat, ReponseEnum
from shared.helpers import convert_date_any_format, relation_entre_table
from utilisateur.models import Utilisateur


@login_required
def liste_sessions(request):

    statut_session = SessionStatut

    context = {
        'statut_session': statut_session
    }

    return render(request, 'sessions/index.html', context)


@login_required
def ajax_datatable_session(request):
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    draw = int(request.GET.get('draw', 1))

    sort_column_index = int(request.GET.get('order[0][column]', 0))
    sort_direction = request.GET.get('order[0][dir]', 'asc')

    search_date_deb = convert_date_any_format(request.GET.get('date_deb', ''))
    search_date_fn = convert_date_any_format(request.GET.get('date_fn', ''))
    search_statut_session = request.GET.get('statut_session', '')

    queryset = Session.objects.filter(deleted_at__isnull=True)

    # Filtres
    if search_date_deb and search_date_fn:
        # Sessions qui chevauchent avec la période recherchée
        queryset = queryset.filter(
            date_debut__lte=search_date_fn,
            date_fin__gte=search_date_deb
        )
    elif search_date_deb:
        queryset = queryset.filter(date_debut__gte=search_date_deb)
    elif search_date_fn:
        queryset = queryset.filter(date_fin__lte=search_date_fn)

    if search_statut_session:
        queryset = queryset.filter(statut_session=search_statut_session)

    # Tri
    sort_columns = {
        0: 'nom',
        1: 'date_publication',
        2: 'date_debut',
        3: 'date_fin',
        4: 'statut_session',
    }

    sort_column = sort_columns.get(sort_column_index, 'id')
    if sort_direction == 'desc':
        sort_column = '-' + sort_column

    queryset = queryset.order_by(sort_column)

    # Pagination
    total_records = queryset.count()

    if length == -1:
        page_queryset = queryset  # pas de pagination
    else:
        page_number = start // length + 1
        paginator = Paginator(queryset, length)
        try:
            page_queryset = paginator.page(page_number)
        except EmptyPage:
            page_queryset = paginator.page(paginator.num_pages)

        # Formatage des données
        data = []
        for sess in page_queryset:
            detail_url = reverse('detail_session', args=[sess.id])
            edit_url = reverse('update_session', args=[sess.id])

            # Bouton "Détails"
            actions_html = f'<a href="{detail_url}"><span class="btn btn-info btn-xs mr-5"><i class="fa fa-eye"></i></span></a>'

            # Bouton "Modifier" et "Supprimer"
            if request.user.is_superadmin or request.user.is_admin:
                actions_html += f'<span class="btn_modifier_session btn btn-warning btn-xs mr-5 mt-3" data-session_id="{sess.id}" data-model_name="session" data-modal_title="Modifier la session" data-href="{edit_url}"><i class="fa fa-edit"></i></span>'

            if request.user.is_superadmin:
                actions_html += f'<span class="btn_supprimer_session btn btn-danger btn-xs mt-3" data-session_id="{sess.id}"><i class="fa fa-trash-o"></i></span>'

            if sess.statut_session:
                statut_html = f'<span class="badge badge-{sess.statut_session.lower().replace(" ", "-")}">{sess.statut_session}</span>'
            else:
                statut_html = '<span class="badge badge-secondary">En attente</span>'

            data.append({
                "id": sess.id,
                "nom": sess.nom if sess.nom else "",
                "date_publication": sess.date_publication.strftime("%d/%m/%Y") if sess.date_publication else "",
                "date_debut": sess.date_debut.strftime("%d/%m/%Y") if sess.date_debut else "",
                "date_fin": sess.date_fin.strftime("%d/%m/%Y") if sess.date_fin else "",
                "nombre_inscrit": '',
                "statut": statut_html,
                "actions": actions_html,
            })

        return JsonResponse({
            "data": data,
            "recordsTotal": total_records,
            "recordsFiltered": total_records,
            "draw": draw,
        })


@login_required
def detail_session(request, session_id):
    session = Session.objects.get(id=session_id)

    if session is None:
        return redirect('sessions')

    statut_session = session.statut_session if session.statut_session else "En attente"
    classe_css_statut = statut_session.lower().replace(' ', '-')
    type_cours = TypeCours.objects.filter(deleted_at__isnull=True).order_by('libelle')
    situation_matrimoniale = SituationMatrimoniale
    genre = Genre
    reponse = ReponseEnum
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')
    certificat_disponible = Certificat.objects.filter(session_id=session.id, date_utilisation__isnull=True, deleted_at__isnull=True).order_by('numero_certificat')
    today = timezone.now().date()

    cheminant_session = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).order_by('-numero_utilisateur')
    nombre_cheminants = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).count()

    cours_qcm = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).exclude(statut_cours=SessionStatut.TERMINE).order_by('-numero_cours')

    now = timezone.now()
    cheminant_mois_en_cours = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True, date_joined__year=now.year, date_joined__month=now.month).count()

    context = {
        'session': session,
        'classe_css_statut': classe_css_statut,
        'type_cours': type_cours,
        'situation_matrimoniale': situation_matrimoniale,
        'genre': genre,
        'reponse': reponse,
        'tribus': tribus,
        'quartiers': quartiers,
        'departements': departements,
        'certificat_disponible': certificat_disponible,
        'cours_qcm': cours_qcm,
        'today': today,
        'nombre_cheminants': nombre_cheminants,
        'cheminant_session': cheminant_session,
        'cheminant_mois_en_cours': cheminant_mois_en_cours,
    }

    return render(request, 'sessions/detail_session.html', context)


@login_required
@transaction.atomic
def add_session(request):
    if request.method == "POST":
        form = SessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.date_publication = timezone.now()
            session.created_at = timezone.now()
            session.created_by = request.user.id
            session.save()  # created_at, updated_at, created_by, updated_by remplis automatiquement

            return JsonResponse({
                'statut': 1,
                'message': "Session enregistrée avec succès !",
                'data': {}
            })

        return JsonResponse({
            'statut': 0,
            'message': "Erreurs de validation",
            'errors': form.errors
        })

    return JsonResponse({'statut': 0, 'message': "Méthode non autorisée"})


@login_required
@transaction.atomic
def update_session(request, session_id):
    try:
        session = Session.objects.get(id=session_id)
    except Session.DoesNotExist:
        return JsonResponse({'statut': 0, 'message': "Session non trouvée"})

    if request.method == 'POST':
        form = SessionForm(request.POST, instance=session)
        if form.is_valid():
            session = form.save(commit=False)
            session.updated_at = timezone.now()
            session.updated_by = request.user.id
            session.save()
            return JsonResponse({
                'statut': 1,
                'message': "Session mise à jour avec succès !",
                'data': {}
            })

        return JsonResponse({
            'statut': 0,
            'message': "Erreurs de validation",
            'errors': form.errors
        })

    # GET : renvoyer le modal
    return render(request, 'partials/update_session_modal.html', {'session': session})


@login_required
def supprimer_session(request):
    if request.method == "POST":

        session_id = request.POST.get('session_id')

        session = Session.objects.get(id=session_id)
        if session.pk is not None:

            if relation_entre_table(session):
                return JsonResponse({
                    'statut': 0,
                    'message': "Impossible de supprimer : cette session est encore utilisée dans une autre table.",
                })

            session.delete()
            response = {
                'statut': 1,
                'message': "Session supprimée avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Session non trouvée !",
            }

        return JsonResponse(response)


@login_required
def certificats_session(request, session_id):
    session = Session.objects.get(id=session_id)

    if session is None:
        return redirect('sessions')

    statut_session = session.statut_session if session.statut_session else "En attente"
    classe_css_statut = statut_session.lower().replace(' ', '-')
    type_cours = TypeCours.objects.filter(deleted_at__isnull=True).order_by('libelle')
    situation_matrimoniale = SituationMatrimoniale
    genre = Genre
    reponse = ReponseEnum
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')

    certificat_session = Certificat.objects.filter(session_id=session.id, deleted_at__isnull=True).order_by('numero_certificat')

    certificat_disponible = Certificat.objects.filter(session_id=session.id, date_utilisation__isnull=True, deleted_at__isnull=True).order_by('numero_certificat')
    today = timezone.now().date()

    cheminant_session = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).order_by('-numero_utilisateur')
    nombre_cheminants = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).count()

    cours_qcm = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).exclude(statut_cours=SessionStatut.TERMINE).order_by('-numero_cours')

    now = timezone.now()
    cheminant_mois_en_cours = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True, date_joined__year=now.year, date_joined__month=now.month).count()

    context = {
        'session': session,
        'classe_css_statut': classe_css_statut,
        'certificat_session': certificat_session,
        'type_cours': type_cours,
        'situation_matrimoniale': situation_matrimoniale,
        'genre': genre,
        'reponse': reponse,
        'tribus': tribus,
        'quartiers': quartiers,
        'departements': departements,
        'certificat_disponible': certificat_disponible,
        'cours_qcm': cours_qcm,
        'today': today,
        'nombre_cheminants': nombre_cheminants,
        'cheminant_session': cheminant_session,
        'cheminant_mois_en_cours': cheminant_mois_en_cours,
    }

    return render(request, 'sessions/certificats.html', context)


@login_required
@transaction.atomic
def add_session_certificat(request, session_id):
    session = Session.objects.get(id=session_id)

    if session is None:
        redirect('detail_session', session_id=session_id)

    if request.method == 'POST':
        form = CertificatForm(request.POST)
        if form.is_valid():
            nombre = form.cleaned_data['nombre']
            date_debut_validite = form.cleaned_data['date_debut_validite']
            date_fin_validite = form.cleaned_data['date_fin_validite']
            session = get_object_or_404(Session, id=session_id)  # Assurez-vous que Session est importé

            # Préfixe basé sur l'année courante (2025)
            annee = timezone.now().year % 100  # Prend les deux derniers chiffres (25)
            prefixe = f"VHCI{annee:02d}"

            # Trouver le dernier numéro utilisé
            last_certificat = Certificat.objects.filter(numero_certificat__startswith=prefixe).order_by('-numero_certificat').first()
            start_num = 1
            if last_certificat:
                last_num = int(last_certificat.numero_certificat.replace(prefixe, '').lstrip('0') or '0')
                start_num = last_num + 1

            # Générer les certificats
            certificats = []
            for i in range(nombre):
                num = start_num + i
                numero_certificat = f"{prefixe}{num:03d}"  # Format avec 5 chiffres (ex: VHCI25001)
                certificats.append(Certificat(
                    id=uuid.uuid4(),
                    numero_certificat=numero_certificat,
                    date_debut_validite=date_debut_validite,
                    date_fin_validite=date_fin_validite,
                    session=session,
                    created_at = timezone.now(),
                    created_by = request.user.id,
                ))

            # Sauvegarder tous les certificats
            Certificat.objects.bulk_create(certificats)

            return JsonResponse({'statut': 1, 'message': f'{nombre} certificats générés avec succès.'})

        return JsonResponse({
            'statut': 0,
            'message': "Erreurs de validation",
            'errors': form.errors
        })
    else:
        redirect('detail_session', session_id=session_id)


@login_required
def detail_session_certificat(request, certificat_id):
    certificat = Certificat.objects.get(id=certificat_id)

    if certificat is None:
        return JsonResponse({
            'statut': 0,
            'message': "Certificat non trouvé",
        })

    cheminant = Utilisateur.objects.filter(certificat_id=certificat_id)

    context = {
        'certificat': certificat,
        'cheminant': cheminant,
    }

    return render(request, 'partials/detail_certificat_modal.html', context)


@login_required
def delete_certificats(request):
    if request.method == 'POST':
        certificats_ids = request.POST.getlist('certificats_ids[]')

        try:
            certificats = Certificat.objects.filter(id__in=certificats_ids)
            if not certificats.exists():
                return JsonResponse({
                    'statut': 0,
                    'message': 'Aucun certificat trouvé.'
                }, status=400)

            supprimes = 0
            ignores = 0

            for certificat in certificats:
                if relation_entre_table(certificat):
                    ignores += 1
                else:
                    certificat.delete()
                    supprimes += 1

            return JsonResponse({
                'statut': 1,
                'message': f'{supprimes} certificats supprimés, {ignores} ignorés car encore utilisés.'
            })

        except Exception as e:
            return JsonResponse({
                'statut': 0,
                'message': f'Erreur lors de la suppression : {str(e)}'
            }, status=500)

    return JsonResponse({
        'statut': 0,
        'message': 'Méthode non autorisée.'
    }, status=405)


@login_required
@transaction.atomic
def add_session_cours(request, session_id):
    session = Session.objects.get(id=session_id)
    if request.method == 'POST':
        # Instancier le formulaire avec les données POST et les fichiers
        form = CoursForm(request.POST, request.FILES)

        if form.is_valid():
            # Préfixe basé sur l'année courante (2025)
            annee = timezone.now().year % 100  # Prend les deux derniers chiffres (25)
            prefixe = f"SCOU{annee:02d}"

            # Trouver le dernier numéro utilisé
            last_cours = Cours.objects.filter(numero_cours__startswith=prefixe).order_by('-numero_cours').first()
            start_num = 1
            if last_cours:
                last_num = int(last_cours.numero_cours.replace(prefixe, '').lstrip('0') or '0')
                start_num = last_num + 1

            # Si le formulaire est valide, les données sont nettoyées
            cours = form.save(commit=False)
            cours.session = session
            cours.numero_cours = f"{prefixe}{start_num:03d}"  # Format avec 5 chiffres (ex: SCOU25001)
            cours.date_publication = timezone.now()
            cours.created_at = timezone.now()
            cours.created_by = request.user.id
            cours.save()
            return JsonResponse({'statut': 1, 'message': 'Cours enregistré avec succès !'})

        return JsonResponse({
            'statut': 0,
            'message': "Erreurs de validation",
            'errors': form.errors
        })
    else:
        # Gérer la requête GET
        return redirect('detail_session', session_id=session_id)


@login_required
def cours_session(request, session_id):
    session = Session.objects.get(id=session_id)

    if session is None:
        return redirect('sessions')

    statut_session = session.statut_session if session.statut_session else "En attente"
    classe_css_statut = statut_session.lower().replace(' ', '-')
    type_cours = TypeCours.objects.filter(deleted_at__isnull=True).order_by('libelle')
    situation_matrimoniale = SituationMatrimoniale
    genre = Genre
    reponse = ReponseEnum
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')

    cours_session = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).order_by('-numero_cours')
    certificat_disponible = Certificat.objects.filter(session_id=session.id, date_utilisation__isnull=True, deleted_at__isnull=True).order_by('numero_certificat')
    today = timezone.now().date()

    cheminant_session = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).order_by('-numero_utilisateur')
    nombre_cheminants = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).count()

    cours_qcm = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).exclude(statut_cours=SessionStatut.TERMINE).order_by('-numero_cours')

    now = timezone.now()
    cheminant_mois_en_cours = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True, date_joined__year=now.year, date_joined__month=now.month).count()

    context = {
        'session': session,
        'classe_css_statut': classe_css_statut,
        'cours_session': cours_session,
        'type_cours': type_cours,
        'situation_matrimoniale': situation_matrimoniale,
        'genre': genre,
        'reponse': reponse,
        'tribus': tribus,
        'quartiers': quartiers,
        'departements': departements,
        'certificat_disponible': certificat_disponible,
        'cours_qcm': cours_qcm,
        'today': today,
        'nombre_cheminants': nombre_cheminants,
        'cheminant_session': cheminant_session,
        'cheminant_mois_en_cours': cheminant_mois_en_cours,
    }

    return render(request, 'sessions/cours.html', context)


@login_required
def detail_session_cours(request, cours_id):
    cours = Cours.objects.get(id=cours_id)

    if cours is None:
        return JsonResponse({
            'statut': 0,
            'message': "Cours non trouvé",
        })

    questions = Question.objects.filter(cours_id=cours_id)

    context = {
        'cours': cours,
        'questions': questions,
    }

    return render(request, 'partials/detail_cours_modal.html', context)


@login_required
@transaction.atomic
def update_session_cours(request, cours_id):
    try:
        cours = Cours.objects.get(id=cours_id)
    except Cours.DoesNotExist:
        return JsonResponse({'statut': 0, 'message': "Cours non trouvé"})

    if request.method == 'POST':
        form = CoursForm(request.POST, request.FILES, instance=cours)
        if form.is_valid():
            # Sauvegarde sans commit pour comparer les fichiers
            new_cours = form.save(commit=False)

            # Vérifier et supprimer les anciens fichiers si remplacés
            if 'cours_video' in request.FILES and cours.cours_video:
                if cours.cours_video.path and os.path.isfile(cours.cours_video.path):
                    cours.cours_video.delete(save=False)

            if 'cours_audio' in request.FILES and cours.cours_audio:
                if cours.cours_audio.path and os.path.isfile(cours.cours_audio.path):
                    cours.cours_audio.delete(save=False)

            if 'cours_texte' in request.FILES and cours.cours_texte:
                if cours.cours_texte.path and os.path.isfile(cours.cours_texte.path):
                    cours.cours_texte.delete(save=False)

            # Mise à jour des métadonnées
            new_cours.updated_at = timezone.now()
            new_cours.updated_by = request.user.id
            new_cours.save()

            return JsonResponse({
                'statut': 1,
                'message': "Cours mis à jour avec succès !",
                'data': {}
            })

        return JsonResponse({
            'statut': 0,
            'message': "Erreurs de validation",
            'errors': form.errors
        })

    # GET : renvoyer le modal
    type_cours = TypeCours.objects.filter(deleted_at__isnull=True).order_by('libelle')
    return render(request, 'partials/update_cours_modal.html', {'cours': cours, 'type_cours': type_cours})


@login_required
def supprimer_cours(request):
    if request.method == "POST":

        cours_id = request.POST.get('cours_id')

        cours = Cours.objects.get(id=cours_id)
        if cours.pk is not None:

            if relation_entre_table(cours):
                return JsonResponse({
                    'statut': 0,
                    'message': "Impossible de supprimer : ce cours est encore utilisé dans une autre table.",
                })

            cours.delete()
            response = {
                'statut': 1,
                'message': "Cours supprimé avec succès !",
            }

        else:
            response = {
                'statut': 0,
                'message': "Cours non trouvé !",
            }

        return JsonResponse(response)


@login_required
def activer_cours(request):
    if request.method == "POST":

        cours_id = request.POST.get('cours_id')

        cours = Cours.objects.get(id=cours_id)
        if cours.pk is not None:
            cours.statut_cours = SessionStatut.ENCOURS
            cours.date_activation = timezone.now()
            cours.save()

            response = {
                'statut': 1,
                'message': "Cours activé avec succès !",
            }

        else:
            response = {
                'statut': 0,
                'message': "Cours non trouvé !",
            }

        return JsonResponse(response)


@login_required
@transaction.atomic
def add_session_cheminant_0(request, session_id):
    if request.method == "POST":
        form = CheminantForm(request.POST, request.FILES)

        if form.is_valid():
            utilisateur = form.save()

            return JsonResponse({
                "statut": 1,
                "message": "Cheminant enregistré avec succès",
            })

        return JsonResponse({
            "statut": 0,
            "message": "Erreurs de validation",
            "errors": form.errors,
        })

    return redirect("detail_session", session_id=session_id)


@login_required
@transaction.atomic
def add_session_cheminant(request, session_id):
    if request.method == "POST":
        print(request.POST, request.FILES)
        form = CheminantForm(request.POST, request.FILES)

        if form.is_valid():
            utilisateur = form.save()

            # Préfixe basé sur l'année courante (2025)
            annee = timezone.now().year % 100  # Prend les deux derniers chiffres (25)
            prefixe = f"CHEM{annee:02d}"

            # Trouver le dernier numéro utilisé
            last_utilisateur = Utilisateur.objects.filter(numero_utilisateur__startswith=prefixe).order_by('-numero_utilisateur').first()
            start_num = 1
            if last_utilisateur:
                last_num = int(last_utilisateur.numero_utilisateur.replace(prefixe, '').lstrip('0') or '0')
                start_num = last_num + 1

            utilisateur.numero_utilisateur = f"{prefixe}{start_num:03d}"
            utilisateur.save()

            # Mettre à jour le certificat comme utilisé
            certificat = Certificat.objects.get(id=utilisateur.certificat_id)
            certificat.date_utilisation = timezone.now()
            certificat.statut_certificat = StatutCertificat.NON_DISPONIBLE
            certificat.save()

            # Créer une inscription après la création de l'utilisateur
            inscription_form = InscriptionForm({
                'utilisateur': utilisateur.id,
                'session': session_id,
                'certificat': form.cleaned_data['certificat_id'].id,
                'statut_inscription': SessionStatut.ENCOURS
            })
            if inscription_form.is_valid():
                inscription_form.save()
                return JsonResponse({
                    "statut": 1,
                    "message": "Cheminant et inscription enregistrés avec succès",
                })
            else:
                print(inscription_form.errors.items())
                return JsonResponse({
                    "statut": 0,
                    "message": "Erreurs lors de l'inscription",
                    "errors": inscription_form.errors,
                })

        # Erreurs de validation → transformer en dict lisible
        print(form.errors.items())
        return JsonResponse({
            "statut": 0,
            "message": "Erreurs de validation",
            "errors": form.errors,
        })

    return redirect("detail_session", session_id=session_id)


@login_required
def cheminant_session(request, session_id):
    session = Session.objects.get(id=session_id)

    if session is None:
        return redirect('sessions')

    statut_session = session.statut_session if session.statut_session else "En attente"
    classe_css_statut = statut_session.lower().replace(' ', '-')
    type_cours = TypeCours.objects.filter(deleted_at__isnull=True).order_by('libelle')
    situation_matrimoniale = SituationMatrimoniale
    genre = Genre
    reponse = ReponseEnum
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')

    cours_session = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).order_by('-numero_cours')
    certificat_disponible = Certificat.objects.filter(session_id=session.id, date_utilisation__isnull=True, deleted_at__isnull=True).order_by('numero_certificat')
    today = timezone.now().date()

    cheminant_session = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).order_by('-numero_utilisateur')
    nombre_cheminants = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).count()

    cours_qcm = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).exclude(statut_cours=SessionStatut.TERMINE).order_by('-numero_cours')

    now = timezone.now()
    cheminant_mois_en_cours = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True, date_joined__year=now.year, date_joined__month=now.month).count()

    context = {
        'session': session,
        'classe_css_statut': classe_css_statut,
        'cours_session': cours_session,
        'type_cours': type_cours,
        'situation_matrimoniale': situation_matrimoniale,
        'genre': genre,
        'reponse': reponse,
        'tribus': tribus,
        'quartiers': quartiers,
        'departements': departements,
        'certificat_disponible': certificat_disponible,
        'cours_qcm': cours_qcm,
        'today': today,
        'nombre_cheminants': nombre_cheminants,
        'cheminant_session': cheminant_session,
        'cheminant_mois_en_cours': cheminant_mois_en_cours,
    }

    return render(request, 'sessions/cheminants.html', context)


@login_required
def detail_session_cheminant(request, utilisateur_id):
    cheminant = Utilisateur.objects.get(id=utilisateur_id)

    if cheminant is None:
        return redirect('detail_session', cheminant.session.id)

    session = Session.objects.filter(id=cheminant.session.id).first()
    situation_matrimoniale = SituationMatrimoniale
    genre = Genre
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')
    certificat_disponible = Certificat.objects.filter(session_id=cheminant.session.id, date_utilisation__isnull=True, deleted_at__isnull=True).order_by('numero_certificat')

    context = {
        'cheminant': cheminant,
        'session': session,
        'situation_matrimoniale': situation_matrimoniale,
        'genre': genre,
        'tribus': tribus,
        'quartiers': quartiers,
        'departements': departements,
        'certificat_disponible': certificat_disponible,
    }

    return render(request, 'sessions/cheminant_detail.html', context)


@login_required
@transaction.atomic
def update_session_cheminants(request, utilisateur_id):
    try:
        utilisateur = Utilisateur.objects.get(id=utilisateur_id)
    except Utilisateur.DoesNotExist:
        return JsonResponse({'statut': 0, 'message': "Cheminant non trouvé"})

    if request.method == 'POST':
        form = CheminantForm(request.POST, request.FILES, instance=utilisateur)
        if form.is_valid():
            utilisateur = form.save()
            return JsonResponse({
                'statut': 1,
                'message': "Cheminant modifié avec succès !"
            })

        print(form.errors.items())
        return JsonResponse({
            'statut': 0,
            'message': "Erreurs de validation",
            'errors': form.errors
        })

    # GET : renvoyer le modal
    cheminant = Utilisateur.objects.get(id=utilisateur_id)
    session = Session.objects.filter(id=utilisateur.session.id).first()
    situation_matrimoniale = SituationMatrimoniale
    genre = Genre
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')
    certificat_disponible = Certificat.objects.filter(session_id=utilisateur.session.id, date_utilisation__isnull=True, deleted_at__isnull=True).order_by('numero_certificat')

    context = {
        'cheminant': cheminant,
        'session': session,
        'situation_matrimoniale': situation_matrimoniale,
        'genre': genre,
        'tribus': tribus,
        'quartiers': quartiers,
        'departements': departements,
        'certificat_disponible': certificat_disponible,
    }
    return render(request, 'partials/update_cheminant_modal.html',context)


@login_required
def export_excel_cheminant_session(request, session_id):
    session = Session.objects.filter(id=session_id).first()
    if not session:
        return redirect('sessions')

    today = timezone.now().date()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CHEMINANTS"

    headers = [
        "NUMERO", "NOM", "PRENOMS", "TELEPHONE", "AUTRE TELEPHONE",
        "SITUATION MATRIMONIALE", "TRIBU", "DEPARTEMENT", "QUARTIER"
    ]

    # Styles
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    bottom_border = Border(bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    # Ligne de titre fusionnée (ligne 1)
    title = f"LISTE DES CHEMINANTS – SESSION {getattr(session, 'nom', session.id)}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    title_cell.border = bottom_border

    # En-têtes (ligne 2)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col_idx)
        c.font = bold_font
        c.border = thin_border
        c.alignment = center_align

    # Données (à partir de la ligne 3)
    cheminants = Utilisateur.objects.order_by("-numero_utilisateur")
    for row_num, ch in enumerate(cheminants, start=3):
        ws.append([
            ch.numero_utilisateur,
            ch.nom, ch.prenoms,  ch.telephone,
            ch.autre_telephone, ch.situation_matrimoniale,
            ch.tribu.libelle if getattr(ch, "tribu", None) else "",
            ch.departement.libelle if getattr(ch, "departement", None) else "",
            ch.quartier.libelle if getattr(ch, "quartier", None) else "",
        ])
        # Bordures fines
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_idx).border = thin_border

    # Ajuster largeur colonnes automatiquement
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 60))

    # Figer les en-têtes (scroll)
    ws.freeze_panes = "A3"

    # Export
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="cheminants-session-{today}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf_cheminant_session(request, session_id):
    session = Session.objects.get(id=session_id)

    if session is None:
        return redirect('sessions')

    pass


@login_required
def import_exemple_cheminant_session(request, session_id):
    session = Session.objects.filter(id=session_id).first()
    if not session:
        return redirect('sessions')

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EXEMPLE_IMPORT"

    headers = [
        "NOM", "PRENOMS", "TELEPHONE", "AUTRE TELEPHONE",
        "SITUATION MATRIMONIALE", "TRIBU", "DEPARTEMENT", "QUARTIER"
    ]
    ws.append(headers)

    # Styles
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Appliquer style au header
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold_font
        c.alignment = center_align
        c.border = thin_border

    # Récupération des valeurs
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')
    situations = [s.value for s in SituationMatrimoniale]

    # Onglet caché "LISTES" pour stocker les valeurs des dropdowns
    ws_lists = wb.create_sheet("LISTES")

    # Remplir les colonnes dans "LISTES"
    for i, val in enumerate(situations, start=1):
        ws_lists.cell(row=i, column=1, value=val)
    for i, val in enumerate([t.libelle for t in tribus], start=1):
        ws_lists.cell(row=i, column=2, value=val)
    for i, val in enumerate([d.libelle for d in departements], start=1):
        ws_lists.cell(row=i, column=3, value=val)
    for i, val in enumerate([q.libelle for q in quartiers], start=1):
        ws_lists.cell(row=i, column=4, value=val)

    # Noms de plages
    ws_lists.title = "LISTES"
    wb.create_named_range("Situations", ws_lists, f"$A$1:$A${len(situations)}")
    wb.create_named_range("Tribus", ws_lists, f"$B$1:$B${len(tribus)}")
    wb.create_named_range("Departements", ws_lists, f"$C$1:$C${len(departements)}")
    wb.create_named_range("Quartiers", ws_lists, f"$D$1:$D${len(quartiers)}")

    # Data Validations (listes déroulantes)
    dv_situation = DataValidation(type="list", formula1="=Situations", allow_blank=True)
    dv_tribu = DataValidation(type="list", formula1="=Tribus", allow_blank=True)
    dv_departement = DataValidation(type="list", formula1="=Departements", allow_blank=True)
    dv_quartier = DataValidation(type="list", formula1="=Quartiers", allow_blank=True)

    ws.add_data_validation(dv_situation)
    ws.add_data_validation(dv_tribu)
    ws.add_data_validation(dv_departement)
    ws.add_data_validation(dv_quartier)

    # Appliquer validations aux colonnes correspondantes (lignes 2 à 500 par ex.)
    dv_situation.add(f"E2:E1000")
    dv_tribu.add(f"F2:F1000")
    dv_departement.add(f"G2:G1000")
    dv_quartier.add(f"H2:H1000")

    # Ajuster largeur colonnes
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx-1]) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    # Masquer la feuille LISTES
    ws_lists.sheet_state = "hidden"

    # Export
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="exemple-import-cheminants.xlsx"'
    wb.save(response)
    return response


@login_required
def supprimer_cheminant(request):
    if request.method == "POST":
        cheminant_id = request.POST.get('cheminant_id')

        try:
            cheminant = Utilisateur.objects.get(id=cheminant_id)
            if cheminant.pk is not None:

                if relation_entre_table(cheminant):
                    return JsonResponse({
                        'statut': 0,
                        'message': "Impossible de supprimer : ce cheminant est encore utilisé dans une autre table.",
                    })

                # Supprimer la photo si elle existe
                if cheminant.photo and default_storage.exists(cheminant.photo.path):
                    try:
                        default_storage.delete(cheminant.photo.path)
                        print(f"Photo supprimée : {cheminant.photo.path}")
                    except Exception as e:
                        print(f"Échec de la suppression de la photo : {e}")

                # Supprimer son inscription
                inscription = Inscription.objects.get(utilisateur_id=cheminant.id)
                inscription.delete()

                # Supprimer son certificat
                certificat = Certificat.objects.get(id=cheminant.certificat.id)
                certificat.delete()

                # Supprimer l'utilisateur
                cheminant.delete()

                response = {
                    'statut': 1,
                    'message': "Cheminant supprimé avec succès !",
                }
            else:
                response = {
                    'statut': 0,
                    'message': "Cheminant non trouvé !",
                }
        except Utilisateur.DoesNotExist:
            response = {
                'statut': 0,
                'message': "Cheminant non trouvé !",
            }

        return JsonResponse(response)


@login_required
@transaction.atomic
def add_qcm_cours_session(request, session_id):
    print('request.POST : ', request.POST)
    if request.method == "POST":
        try:
            cours_id = request.POST.get("cours_id")
            if not cours_id:
                return JsonResponse({"statut": 0, "errors": {"cours_id": ["Cours obligatoire"]}})

            cours = Cours.objects.get(id=cours_id)

            # Extraire toutes les questions en gérant l'encodage URL
            questions_data = {}

            # Parcourir tous les éléments POST pour identifier les questions
            for key, value_list in request.POST.lists():
                # Décoder l'URL encoding
                decoded_key = urllib.parse.unquote(key)

                # Vérifier si c'est une question
                if decoded_key.startswith("questions[") and decoded_key.endswith("]"):
                    # Extraire l'index de la question
                    question_index = decoded_key.split("[")[1].split("]")[0]

                    # Initialiser la structure pour cette question si nécessaire
                    if question_index not in questions_data:
                        questions_data[question_index] = {
                            'libelle': '',
                            'reponses': [],
                            'types': []
                        }

                    questions_data[question_index]['libelle'] = value_list[0] if value_list else ''

                # Vérifier si ce sont des réponses
                elif decoded_key.startswith("reponses[") and "[]" in decoded_key:
                    # Extraire l'index de la question (format: reponses[X][])
                    question_index = decoded_key.split("[")[1].split("]")[0]

                    if question_index not in questions_data:
                        questions_data[question_index] = {
                            'libelle': '',
                            'reponses': [],
                            'types': []
                        }

                    questions_data[question_index]['reponses'] = value_list

                # Vérifier si ce sont des types de réponses
                elif decoded_key.startswith("type_reponses[") and "[]" in decoded_key:
                    question_index = decoded_key.split("[")[1].split("]")[0]

                    if question_index not in questions_data:
                        questions_data[question_index] = {
                            'libelle': '',
                            'reponses': [],
                            'types': []
                        }

                    questions_data[question_index]['types'] = value_list

            # Traitement des questions
            for question_index, question_data in questions_data.items():
                question_libelle = question_data['libelle']

                if not question_libelle.strip():
                    continue

                reponses = question_data['reponses']
                types = question_data['types']

                # Vérifier que les listes ont la même longueur
                max_length = max(len(reponses), len(types))

                # Étendre les listes si nécessaire
                reponses.extend([''] * (max_length - len(reponses)))
                types.extend(['Faux'] * (max_length - len(types)))

                # Création de la question
                question = Question.objects.create(
                    libelle=question_libelle.strip(),
                    date_publication=timezone.now(),
                    created_at = timezone.now(),
                    created_by = request.user.id,
                    cours=cours,
                )

                # Création des réponses liées
                for libelle, statut in zip(reponses, types):
                    if not libelle.strip():
                        continue

                    reponse = Reponse.objects.create(
                        libelle=libelle.strip(),
                        date_publication=timezone.now(),
                        created_at = timezone.now(),
                        created_by = request.user.id,
                        question=question,
                        statut_reponse=statut or "Faux",
                    )

            return JsonResponse({"statut": 1, "message": "QCM enregistré avec succès !"})

        except Exception as e:
            print(f'Erreur lors de l\'enregistrement: {e}')
            return JsonResponse({"statut": 0, "message": f"Erreur interne : {e}"})


@login_required
def qcm_cours_session(request, session_id):
    session = Session.objects.get(id=session_id)

    if session is None:
        return redirect('sessions')

    statut_session = session.statut_session if session.statut_session else "En attente"
    classe_css_statut = statut_session.lower().replace(' ', '-')
    type_cours = TypeCours.objects.filter(deleted_at__isnull=True).order_by('libelle')
    situation_matrimoniale = SituationMatrimoniale
    genre = Genre
    reponse = ReponseEnum
    tribus = Tribu.objects.filter(deleted_at__isnull=True).order_by('libelle')
    quartiers = Quartier.objects.filter(deleted_at__isnull=True).order_by('libelle')
    departements = Departement.objects.filter(deleted_at__isnull=True).order_by('libelle')

    cours_session = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).order_by('-numero_cours')
    certificat_disponible = Certificat.objects.filter(session_id=session.id, date_utilisation__isnull=True, deleted_at__isnull=True).order_by('numero_certificat')
    today = timezone.now().date()

    cheminant_session = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).order_by('-numero_utilisateur')
    nombre_cheminants = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True).count()

    cours_qcm = Cours.objects.filter(session_id=session.id, deleted_at__isnull=True).exclude(statut_cours=SessionStatut.TERMINE).order_by('-numero_cours')

    now = timezone.now()
    cheminant_mois_en_cours = Utilisateur.objects.filter(session_id=session.id, user_etudiant=True, date_joined__year=now.year, date_joined__month=now.month).count()

    context = {
        'session': session,
        'classe_css_statut': classe_css_statut,
        'cours_session': cours_session,
        'type_cours': type_cours,
        'situation_matrimoniale': situation_matrimoniale,
        'genre': genre,
        'reponse': reponse,
        'tribus': tribus,
        'quartiers': quartiers,
        'departements': departements,
        'certificat_disponible': certificat_disponible,
        'cours_qcm': cours_qcm,
        'today': today,
        'nombre_cheminants': nombre_cheminants,
        'cheminant_session': cheminant_session,
        'cheminant_mois_en_cours': cheminant_mois_en_cours,
    }

    return render(request, 'sessions/qcm_cours_session.html', context)


@login_required
def ajax_datatable_qcm_cours_session(request):
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    draw = int(request.GET.get('draw', 1))

    sort_column_index = int(request.GET.get('order[0][column]', 0))
    sort_direction = request.GET.get('order[0][dir]', 'asc')

    search_cours = request.GET.get('cours_id', '')

    queryset = Question.objects.filter(deleted_at__isnull=True)

    # Filtres
    if search_cours:
        queryset = queryset.filter(cours_id=search_cours)

    # Tri
    sort_columns = {
        0: 'libelle',
        2: 'point',
        1: 'date_publication',
        3: 'statut_question',
        4: 'cours_id',
    }

    sort_column = sort_columns.get(sort_column_index, 'id')
    if sort_direction == 'desc':
        sort_column = '-' + sort_column

    queryset = queryset.order_by(sort_column)

    # Pagination
    total_records = queryset.count()

    if length == -1:
        page_queryset = queryset  # pas de pagination
    else:
        page_number = start // length + 1
        paginator = Paginator(queryset, length)
        try:
            page_queryset = paginator.page(page_number)
        except EmptyPage:
            page_queryset = paginator.page(paginator.num_pages)

        # Formatage des données
        data = []
        for qcm in page_queryset:
            detail_url = reverse('detail_session_cours_qcm', args=[qcm.id])
            edit_url = reverse('update_session_cours_qcm', args=[qcm.id])

            # Bouton "Détails"
            actions_html = f'<span class="btn_detail_qcm btn btn-info btn-xs mr-5 text-center" data-qcm_cours_id="{qcm.id}" data-model_name="qcm_cours" data-modal_title="Détail du QCM" data-href="{detail_url}"><i class="fa fa-eye"></i></span>'

            # Bouton "Modifier"
            if request.user.is_superadmin:
                actions_html += f'<span class="btn_modifier_qcm_cours btn btn-warning btn-xs mr-5 text-center" data-qcm_cours_id="{qcm.id}" data-model_name="qcm_cours" data-modal_title="Modifier le QCM" data-href="{edit_url}"><i class="fa fa-edit"></i></span>'

            # Bouton "Supprimer"
            if request.user.is_superadmin:
                actions_html += f'<span class="btn_supprimer_qcm_cours btn btn-danger btn-xs text-center" data-qcm_cours_id="{qcm.id}"><i class="fa fa-trash-o"></i></span>'

            if qcm.cours_id:
                statut_html = f'<span class="badge badge-{qcm.statut_question.lower().replace(" ", "-")}">{qcm.statut_question}</span>'
            else:
                statut_html = '<span class="badge badge-secondary">En attente</span>'

            reponses = ''

            data.append({
                "id": qcm.id,
                "libelle": qcm.libelle if qcm.libelle else "",
                "total_point": qcm.point if qcm.point else "",
                "date_publication": qcm.date_publication.strftime("%d/%m/%Y") if qcm.date_publication else "",
                "reponses": reponses,
                "statut": statut_html,
                "actions": actions_html,
            })

        return JsonResponse({
            "data": data,
            "recordsTotal": total_records,
            "recordsFiltered": total_records,
            "draw": draw,
        })


@login_required
def detail_session_cours_qcm(request, qcm_id):
    qcm = Question.objects.get(id=qcm_id)

    if qcm is None:
        return JsonResponse({
            'statut': 0,
            'message': "QCM non trouvé",
        })

    reponses = Reponse.objects.filter(question_id=qcm.id)

    context = {
        'qcm': qcm,
        'reponses': reponses,
    }

    return render(request, 'partials/detail_question_reponse_modal.html', context)


@login_required
@transaction.atomic
def update_session_cours_qcm(request, qcm_id):
    try:
        qcm = Question.objects.get(id=qcm_id)
    except Cours.DoesNotExist:
        return JsonResponse({'statut': 0, 'message': "Question non trouvée"})

    if request.method == 'POST':
        cours_id = request.POST.get("cours_id")
        if not cours_id:
            return JsonResponse({"statut": 0, "errors": {"cours_id": ["Cours obligatoire"]}})

        cours = Cours.objects.get(id=cours_id)

        # Récupérer les listes
        reponses = request.POST.getlist("reponses[]")
        types = request.POST.getlist("type_reponses[]")
        points = request.POST.getlist("points[]")

        # Mise à jour de la question
        qcm.cours = cours
        qcm.libelle = request.POST.get("question")
        qcm.date_publication = timezone.now()

        # recalcul du total des points
        qcm.point = sum(int(p) for t, p in zip(types, points) if t == "Vrai" and p.isdigit())
        qcm.save()

        # Supprimer les anciennes réponses avant de recréer
        Reponse.objects.filter(question=qcm).delete()

        # Réinsertion des réponses
        for libelle, statut, point in zip(reponses, types, points):
            if not libelle.strip():
                continue
            if not point.strip():
                continue
            Reponse.objects.create(
                libelle=libelle.strip(),
                point=point.strip(),
                date_publication=timezone.now(),
                created_at = timezone.now(),
                created_by = request.user.id,
                question=qcm,
                statut_reponse=statut or "Faux",
            )

        return JsonResponse({"statut": 1, "message": "QCM mis à jour avec succès !"})

    # GET : renvoyer le modal
    reponse = ReponseEnum

    reponses = Reponse.objects.filter(question_id=qcm.id, deleted_at__isnull=True).order_by('created_at')

    context = {
        "qcm":qcm,
        "reponse":reponse,
        "reponses":reponses,
    }
    return render(request, 'partials/update_question_reponse_modal.html', context)


@login_required
def supprimer_qcm_cours_session(request):
    if request.method == "POST":
        qcm_cours_id = request.POST.get('qcm_cours_id')

        try:
            qcm = Question.objects.get(id=qcm_cours_id)
            if qcm.pk is not None:

                if relation_entre_table(qcm):
                    return JsonResponse({
                        'statut': 0,
                        'message': "Impossible de supprimer : ce QCM est encore utilisé dans une autre table.",
                    })

                # Supprimer ses réponses
                reponse = Reponse.objects.filter(question_id=qcm.id)
                reponse.delete()

                # Supprimer la question
                qcm.delete()

                response = {
                    'statut': 1,
                    'message': "Question supprimée avec succès !",
                }
            else:
                response = {
                    'statut': 0,
                    'message': "Question non trouvée !",
                }
        except Question.DoesNotExist:
            response = {
                'statut': 0,
                'message': "Question non trouvée !",
            }

        return JsonResponse(response)
